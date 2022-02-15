import pdfplumber
import os
import re
import datetime
from openpyxl import load_workbook

wb = load_workbook('修改文件名.xlsx')
wb1 = wb['简称基础数据']
wb2 = wb1.max_row

def get_abbr():
	dict1 = {}
	for i in range(1,wb2):
		dict1[wb1['A' + str(i + 1)].value] = wb1['B' + str(i + 1)].value
	return dict1

pdf_lst = [f for f in os.listdir() if f.endswith('.PDF')]

def get_pdf_data():
	list1 = []
	for pdf in pdf_lst:
		with pdfplumber.open(pdf) as pdf:
			text = pdf.pages[-1].extract_text()
			text1 = pdf.pages[0].extract_text()
			text2 = pdf.pages[len(pdf.pages)-2].extract_text()
			if len(pdf.pages) == 1:
				data1 = re.findall(r'发票类型：.*\n发票日期：(.+)\n客户：(.+)\s发票号码：(\d+)\n{1,3}.+\d+\.\d{2}.*(人民币|欧元|美元)\s\d+\.\d{2}.+合计\s(\d+\.\d{2})',text1,re.DOTALL)
				data2 = ""
			else:
				if "合计" in text:
					data1 = re.findall(r'发票类型：.*\n发票日期：(.+)\n客户：(.+)\s发票号码：(\d+)\n{1,3}.+\d+\.\d{2}.*(人民币|欧元|美元)\s\d+\.\d{2}',text1,re.DOTALL)
					data2 = re.findall(r'合计\s(\d+\.\d{2})',text,re.DOTALL)	
				else:
					data1 = re.findall(r'发票类型：.*\n发票日期：(.+)\n客户：(.+)\s发票号码：(\d+)\n{1,3}.+\d+\.\d{2}.*(人民币|欧元|美元)\s\d+\.\d{2}',text1,re.DOTALL)
					data2 = re.findall(r'合计\s(\d+\.\d{2})',text2,re.DOTALL)
			for i in data1:
				i += tuple(data2)
				list1.append(i)
	return list1

currency_dict = {
	"人民币":"￥",
	"美元":"$",
	"欧元":"€"
}


for b, c in zip(pdf_lst,get_pdf_data()):
	os.rename(os.getcwd() + "\\" + b, os.getcwd() + "\\" + datetime.datetime.strptime(c[0],'%Y/%m/%d').strftime('%Y年%m月%d日')  + c[2]  + get_abbr().get(c[1],"【缺失简称】")  + currency_dict[c[3]] + c[4].rstrip('0').rstrip('.') + "账单.PDF")